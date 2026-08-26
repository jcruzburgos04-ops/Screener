#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
SCREENER LOCAL TIPO FINVIZ  ·  ASH (diario + semanal), EMAs, ADR%, RSI, volumen
================================================================================

Motor: Python calcula -> Excel filtra. Gratis, sobre yfinance.

El ASH es una traduccion linea por linea de
"Absolute Strength Histogram v2 | jh" (Pine v4, original de alexgrover),
con una sola diferencia deliberada: el Pine plotea abs(SmthBulls - SmthBears)
y le pone el signo por color. Aca la columna guarda la DIFERENCIA CON SIGNO:

    ASH = SmthBulls - SmthBears

    > 0  la linea verde (bulls) esta por encima de la roja
    < 0  la verde esta por debajo

USO
---
    pip install yfinance pandas numpy openpyxl
    python screener.py                         # corrida real
    python screener.py --sin-fundamentales     # sin float/sector, mas rapido
    python screener.py --universo mi.csv --out salida.xlsx --periodo 3y

DONDE TOCAR
-----------
    CFG_ASH   parametros del indicador (iguales a los del Pine)
    FILTROS   el screen. None desactiva el filtro.
================================================================================
"""

import argparse
import json
import sys
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import pandas as pd

# ==============================================================================
# 1. CONFIGURACION
# ==============================================================================

BENCHMARK = "SPY"
PERIODO = "3y"             # 3y da ~150 barras semanales, comodo para el ASH W
MIN_BARRAS = 220           # barras diarias minimas para aceptar un simbolo
MIN_BARRAS_SEM = 30        # barras semanales minimas para calcular el ASH W

# ---- ASH: mismos nombres y defaults que el Pine ------------------------------
#   modo:    "RSI" | "STOCHASTIC" | "ADX"
#   ma_type: "ALMA" | "EMA" | "WMA" | "SMA" | "SMMA" | "HMA"
CFG_ASH = {
    "length": 16,
    "smooth": 4,
    "modo": "RSI",
    "ma_type": "EMA",
    "alma_offset": 0.85,
    "alma_sigma": 6.0,
}

# ---- Otros indicadores -------------------------------------------------------
PARAGON = {"rapida": 100, "lenta": 200, "k": 2, "fresco": 5,
           "rv_len": 365, "rv_fuente": "hl2"}
#   k = velas de 4h por rueda. Con 2 (una rueda de 6,5 h son dos velas de 4 h)
#   el par 100/200 de 4h equivale a 50/100 en diario. Con 6 (cripto) da 17/33,
#   que son los numeros que documenta el Pine original.
ADR_LEN = 20               # ventana del Average Daily Range
RSI_LEN = 14
ATR_LEN = 14
ADX_LEN = 14

# ---- EL SCREEN. None = filtro apagado ----------------------------------------
FILTROS = {
    # --- Volumen y liquidez ---
    "vol20_min":           300_000,      # acciones promedio 20 ruedas
    "dolar_vol20_min":     5_000_000,    # volumen en dolares promedio 20 ruedas
    "rel_vol_min":         None,         # volumen de hoy / promedio 20d

    # --- Movimiento ---
    "adr_pct_min":         None,         # ADR% >= X  (3.0 = se mueve 3% por dia)
    "adr_pct_max":         None,
    "atr_pct_max":         None,

    # --- RSI ---
    "rsi14_min":           None,
    "rsi14_max":           None,

    # --- EMAs ---
    "par_b_sesgo":         None,         # True / False / None
    "par_a_sesgo":         None,
    "regimen_ord":         None,

    # --- ASH diario ---
    "ash_d_min":           None,         # valor de la diferencia
    "ash_d_positivo":      None,         # True = verde por encima de roja
    "ash_d_creciendo":     None,         # True = la diferencia se agranda
    "ash_d_barras_cruce_max": None,      # cruce alcista de hace <= N ruedas

    # --- ASH semanal ---
    "ash_w_min":           None,
    "ash_w_positivo":      None,
    "ash_w_creciendo":     None,

    # --- Precio, float, tendencia ---
    "precio_min":          5.0,
    "precio_max":          None,
    "float_musd_min":      None,
    "float_musd_max":      None,
    "float_shares_max_m":  None,
    "rotacion_float_min":  None,
    "adx14_min":           None,
    "dist_max52s_max":     None,         # % maximo por debajo del maximo de 52s
    "rs_rank_min":         None,         # 1..99
}

# Tabla de reglas: (clave_filtro, columna, operador)
REGLAS = [
    ("vol20_min",          "vol20",           "ge"),
    ("dolar_vol20_min",    "dolar_vol20",     "ge"),
    ("rel_vol_min",        "rel_vol",         "ge"),
    ("adr_pct_min",        "adr_pct",         "ge"),
    ("adr_pct_max",        "adr_pct",         "le"),
    ("atr_pct_max",        "atr_pct",         "le"),
    ("rsi14_min",          "rsi14",           "ge"),
    ("rsi14_max",          "rsi14",           "le"),
    ("par_b_sesgo",        "par_b_sesgo",     "bool"),
    ("par_a_sesgo",        "par_a_sesgo",     "bool"),
    ("ash_d_min",          "ash_d",           "ge"),
    ("ash_d_positivo",     "ash_d_positivo",  "bool"),
    ("ash_d_creciendo",    "ash_d_creciendo", "bool"),
    ("ash_w_min",          "ash_w",           "ge"),
    ("ash_w_positivo",     "ash_w_positivo",  "bool"),
    ("ash_w_creciendo",    "ash_w_creciendo", "bool"),
    ("precio_min",         "precio",          "ge"),
    ("precio_max",         "precio",          "le"),
    ("float_musd_min",     "float_musd",      "ge"),
    ("float_musd_max",     "float_musd",      "le"),
    ("float_shares_max_m", "float_shares_m",  "le"),
    ("rotacion_float_min", "rotacion_float",  "ge"),
    ("adx14_min",          "adx14",           "ge"),
    ("rs_rank_min",        "rs_rank",         "ge"),
]

# Columnas del Excel: (clave, titulo, formato)
COLUMNAS = [
    ("ticker",            "Ticker",          "@"),
    ("local",             "Local/CEDEAR",    "@"),
    ("grupo",             "Grupo",           "@"),
    ("nombre",            "Nombre",          "@"),
    ("sector",            "Sector",          "@"),
    ("industria",         "Industria",       "@"),
    ("pais",              "Pais",            "@"),
    ("precio",            "Precio",          "#,##0.00"),
    ("chg_pct",           "Chg %",           "0.0%"),
    # --- ASH ---
    ("ash_d",             "ASH D",           "#,##0.0000"),
    ("ash_d_pend",        "ASH D pend",      "#,##0.0000"),
    ("ash_d_norm",        "ASH D norm",      "0.000"),
    ("ash_d_barras_cruce", "ASH D barras",   "0"),
    ("ash_w",             "ASH W",           "#,##0.0000"),
    ("ash_w_pend",        "ASH W pend",      "#,##0.0000"),
    ("ash_w_norm",        "ASH W norm",      "0.000"),
    ("ash_bulls_d",       "Bulls D",         "#,##0.0000"),
    ("ash_bears_d",       "Bears D",         "#,##0.0000"),
    # --- Momentum / volatilidad ---
    ("rsi14",             "RSI 14",          "0.0"),
    ("adr_pct",           "ADR %",           "0.00"),
    ("atr_pct",           "ATR %",           "0.00"),
    ("adx14",             "ADX 14",          "0.0"),
    # --- EMAs ---
    ("regimen",           "Régimen",         "@"),
    ("par_b_pos",         "Precio vs B",     "@"),
    ("par_b_dist",        "Dist. nube B",    "0.0%"),
    ("par_b_ancho",       "Ancho B",         "0.0%"),
    ("par_b_cruce",       "Cruce B",         "0"),
    ("par_a_pos",         "Precio vs A",     "@"),
    ("par_a_dist",        "Dist. nube A",    "0.0%"),
    ("par_a_cruce",       "Cruce A",         "0"),
    ("vs_rvwap",          "vs rVWAP 365",    "0.0%"),
    # --- Performance ---
    ("perf_1s",           "1 sem",           "0.0%"),
    ("perf_1m",           "1 mes",           "0.0%"),
    ("perf_3m",           "3 meses",         "0.0%"),
    ("perf_6m",           "6 meses",         "0.0%"),
    ("perf_12m",          "12 meses",        "0.0%"),
    ("rs_3m",             "RS 3m vs bench",  "0.0%"),
    ("rs_rank",           "RS rank",         "0"),
    ("dist_max52s",       "Desde max 52s %", "0.0%"),
    ("dist_min52s",       "Sobre min 52s %", "0.0%"),
    # --- Volumen y float ---
    ("volumen",           "Volumen",         "#,##0"),
    ("vol20",             "Vol prom 20d",    "#,##0"),
    ("rel_vol",           "Rel volumen",     "0.00"),
    ("dolar_vol20",       "$ Vol 20d",       "#,##0"),
    ("float_shares_m",    "Float (M acc)",   "#,##0.0"),
    ("float_musd",        "Float ($M)",      "#,##0"),
    ("rotacion_float",    "Rotacion float",  "0.000"),
    ("mcap_musd",         "Market cap ($M)", "#,##0"),
]

MAPA_CEDEARS = Path("cedears.csv")   # local -> subyacente que cotiza afuera

CACHE_PRECIOS = Path("cache_precios.pkl")   # lo reusa la app web
CACHE_META = Path("cache_fundamentales.json")
CACHE_DIAS = 7

CUARENTENA = Path("sin_datos.json")   # lo que Yahoo no devuelve, con su castigo
CUARENTENA_FALLOS = 3                 # fallos seguidos para dejar de pedirlo
CUARENTENA_DIAS = 7                   # cuanto dura el castigo antes de reintentar
MAX_INDIVIDUALES = 150                # tope de reintentos de a uno, por tiempo


# ==============================================================================
# 2. MEDIAS MOVILES (las seis del Pine)
# ==============================================================================

def sma(s, n):
    return s.rolling(int(n), min_periods=int(n)).mean()


def ema(s, n):
    return s.ewm(span=int(n), adjust=False).mean()


# ---------------------------------------------------------------------------
# PARAGON  ·  EMA 100/200 ancladas, y rVWAP 365d
#
# Reconstruccion de los dos indicadores privados de DocXBT (The Paragon Group).
# Los dos son EL MISMO par 100/200; lo unico que cambia es el timeframe de
# anclaje: "Paragon Daily" ancla a 4h y "Paragon Weekly" ancla a 1D.
# ---------------------------------------------------------------------------

def ema_pine(s, n):
    """
    ta.ema de Pine Script, que NO es lo mismo que ewm() de pandas.

    Diferencias que importan, porque el usuario cruza estos numeros contra
    TradingView:
      - la semilla es la SMA de los primeros n valores, no el primer precio
        (que es lo que hace ewm(adjust=False)) ni un promedio expansivo
        (ewm(adjust=True));
      - devuelve NaN durante las primeras n-1 velas, sin rellenar hacia atras.

    OJO: la ema() de mas arriba NO se toca. Esa la consume el ASH, y su
    semilla "primer valor" es justo lo que hace que la paridad Python <-> JS
    del ASH de 6,7e-14 se sostenga. Son dos medias distintas a proposito.
    """
    n = int(n)
    x = np.asarray(s, dtype=float)
    out = np.full(len(x), np.nan)
    if len(x) < n or n < 1:
        return pd.Series(out, index=s.index)
    a = 2.0 / (n + 1.0)
    p = float(np.mean(x[:n]))          # semilla = SMA de los primeros n
    out[n - 1] = p
    for i in range(n, len(x)):
        v = x[i]
        if not np.isfinite(v):
            out[i] = p
            continue
        p = a * v + (1.0 - a) * p
        out[i] = p
    return pd.Series(out, index=s.index)


def largo_equivalente(largo, k):
    """
    Convierte la longitud de una EMA del timeframe ancla al del grafico.

    NO es largo/k. Lo que se conserva es la tasa de decaimiento por unidad de
    tiempo calendario, y eso es multiplicativo:

        (1 - a_ancla)^k = (1 - a_destino)   ->   a_destino = 1 - (1 - a_ancla)^k
        L_destino = 2/a_destino - 1

    Para k chico las dos formulas casi coinciden (EMA 200 de 4h en diario:
    33,33 lineal contra 33,39 exacta), pero para k grande divergen feo: en
    semanal la EMA 100 de 4h da 2,38 lineal y 2,52 exacta, y eso cambia el
    redondeo de 2 a 3.

    Con k=6 (cripto, seis velas de 4h por dia) da 17 y 33, que son exactamente
    los numeros que documenta el Pine original. Es la comprobacion de que la
    formula esta bien.
    """
    a = 2.0 / (float(largo) + 1.0)
    a_dest = 1.0 - (1.0 - a) ** float(k)
    if a_dest <= 0:
        return int(largo)
    return max(2, int(round(2.0 / a_dest - 1.0)))


def rvwap_expansivo(df, n=365, fuente="hl2"):
    """
    VWAP rolling sobre las ultimas min(t+1, n) velas diarias.

    El min() es deliberado: mientras el simbolo tenga menos de n velas la
    ventana arranca desde la primera disponible en vez de devolver NaN. O sea
    que hasta la vela n el valor es un VWAP anclado al inicio del historico, y
    de ahi en adelante pasa a ser la ventana movil. La transicion es continua,
    sin salto.

    Devuelve (serie, ventana_llena_bool).

    Fuente hl2 por defecto, que es la del Pine del usuario. El VWAP nativo de
    TradingView usa hlc3, asi que queda configurable.
    """
    n = int(n)
    if fuente == "hlc3":
        px = (df["High"] + df["Low"] + df["Close"]) / 3.0
    elif fuente == "close":
        px = df["Close"].astype(float)
    else:
        px = (df["High"] + df["Low"]) / 2.0
    vol = df["Volume"].astype(float).fillna(0.0)
    pv = (px * vol).cumsum()
    cv = vol.cumsum()
    # la diferencia de acumulados: donde todavia no hay n velas, el desplazado
    # es NaN y el fillna(0) deja el acumulado COMPLETO, que es justo la ventana
    # expansiva que se busca
    pvn = pv - pv.shift(n).fillna(0.0)
    cvn = cv - cv.shift(n).fillna(0.0)
    out = pd.Series(np.where(cvn > 0, pvn / cvn, np.nan), index=df.index)
    return out, len(df) >= n


def paragon_conjunto(cierre, rapida=100, lenta=200, k=1):
    """
    Un conjunto Paragon (el par 100/200) llevado al timeframe de las barras
    que se le pasan.

    k = velas del ancla por vela de la serie. Con k=1 el ancla ES la serie y
    las longitudes van tal cual (el caso del conjunto SEMANAL sobre diarias:
    exacto, sin aproximar nada). Con k>1 el ancla es mas fina que la serie y
    las longitudes se convierten con largo_equivalente(), que es una
    aproximacion: misma tasa de decaimiento, pero las dos EMAs comen series
    distintas, asi que los valores no coinciden con los del ancla de verdad.

    Devuelve (rapida_serie, lenta_serie, largo_rapida_usado, largo_lenta_usado).
    """
    lr = largo_equivalente(rapida, k) if k != 1 else int(rapida)
    ll = largo_equivalente(lenta, k) if k != 1 else int(lenta)
    a, b = ema_pine(cierre, lr), ema_pine(cierre, ll)
    # WARMUP: la longitud convertida gobierna la FORMA de la curva, pero no
    # cuando puede existir. Para que el conjunto imprima hace falta juntar
    # `largo` velas del ancla, o sea ceil(largo/k) velas de la serie. Con
    # k=6 la EMA 200 de 4h necesita 200/6 = 33,33 dias, asi que la primera
    # vela diaria que las completa es la 34 -- y no la 33, que es donde
    # imprimiria una EMA(33) suelta. Ese uno de diferencia es justo el dato
    # que identifica al indicador de Doc, asi que se respeta.
    a = _recortar_warmup(a, int(np.ceil(rapida / float(k))))
    b = _recortar_warmup(b, int(np.ceil(lenta / float(k))))
    return a, b, lr, ll


def _recortar_warmup(s, velas):
    """Deja en NaN las primeras `velas - 1`, sin tocar el resto."""
    if velas <= 1:
        return s
    out = s.copy()
    out.iloc[:min(velas - 1, len(out))] = np.nan
    return out


def senales_paragon(df, rap, len_, atr=None):
    """
    Las columnas derivadas de un conjunto: sesgo, posicion del precio respecto
    de la nube, ancho, distancia al borde mas cercano y velas desde el cruce.

    Todo NaN/None si el conjunto todavia no imprimio (simbolo joven): no se
    inventa nada, se marca y se filtra.
    """
    px = float(df["Close"].iloc[-1])
    a, b = rap.iloc[-1], len_.iloc[-1]
    if not (np.isfinite(a) and np.isfinite(b)):
        return {"sesgo": None, "pos": "", "ancho": np.nan,
                "dist": np.nan, "dist_atr": np.nan, "cruce": np.nan}
    a, b = float(a), float(b)
    techo, piso = max(a, b), min(a, b)
    pos = "arriba" if px > techo else ("abajo" if px < piso else "adentro")
    borde = techo if px > techo else (piso if px < piso else
                                      (techo if techo - px < px - piso else piso))
    dist = (px - borde) / px
    val = rap - len_
    v = val.dropna()
    cruce = np.nan
    if len(v) >= 2:
        signo = np.sign(v.to_numpy())
        ult = signo[-1]
        cruce = 0
        for i in range(len(signo) - 2, -1, -1):
            if signo[i] != ult and signo[i] != 0:
                break
            cruce += 1
        if cruce >= len(signo) - 1:
            cruce = np.nan          # nunca cruzo en el historial disponible
    return {"sesgo": bool(a >= b), "pos": pos,
            "ancho": abs(a - b) / px, "dist": dist,
            "dist_atr": (px - borde) / atr if (atr and atr == atr and atr > 0)
                        else np.nan,
            "cruce": float(cruce) if cruce == cruce else np.nan}


def _conv(s, pesos):
    """Media movil por convolucion: mismo resultado que un rolling, mas rapido."""
    n = len(pesos)
    x = s.to_numpy(dtype=float)
    out = np.full(len(x), np.nan)
    if len(x) >= n:
        out[n - 1:] = np.convolve(x, pesos[::-1], mode="valid")
    return pd.Series(out, index=s.index)


def wma(s, n):
    n = int(n)
    w = np.arange(1.0, n + 1.0)
    return _conv(s, w / w.sum())


def alma(s, n, offset=0.85, sigma=6.0):
    n = int(n)
    m = offset * (n - 1)
    d = n / float(sigma)
    i = np.arange(n, dtype=float)
    w = np.exp(-((i - m) ** 2) / (2 * d * d))
    return _conv(s, w / w.sum())


def hma(s, n):
    n = int(n)
    return wma(2 * wma(s, n // 2) - wma(s, n), int(round(np.sqrt(n))))


def smma(s, n):
    """
    Replica literal del SMMA del Pine de referencia:

        w = wma(src, len)
        result := na(w[1]) ? sma(src, len) : (w[1]*(len-1) + src) / len

    Ojo: NO es la SMMA recursiva clasica (Wilder). El `result` del Pine no se
    referencia a si mismo, usa la WMA de la barra anterior. Lo dejo igual para
    que los numeros coincidan con el indicador del grafico.
    """
    n = int(n)
    w = wma(s, n)
    r = (w.shift(1) * (n - 1) + s) / n
    return r.fillna(sma(s, n))


def rma(s, n):
    """Media de Wilder. No esta en el Pine del ASH, pero RSI/ADX/ATR la usan."""
    return s.ewm(alpha=1.0 / int(n), adjust=False).mean()


def ma(tipo, s, n, alma_offset=0.85, alma_sigma=6.0):
    t = tipo.upper()
    if t == "SMA":
        return sma(s, n)
    if t == "EMA":
        return ema(s, n)
    if t == "WMA":
        return wma(s, n)
    if t == "SMMA":
        return smma(s, n)
    if t == "HMA":
        return hma(s, n)
    if t == "ALMA":
        return alma(s, n, alma_offset, alma_sigma)
    raise ValueError(f"MA desconocida: {tipo}")


# ==============================================================================
# 3. ASH
# ==============================================================================

def calc_ash(df, length=9, smooth=3, modo="RSI", ma_type="WMA",
             alma_offset=0.85, alma_sigma=6.0, src="Close"):
    """
    Absolute Strength Histogram v2 (jh / alexgrover), traducido de Pine v4.

    Devuelve (SmthBulls, SmthBears, SmthBulls - SmthBears).
    El tercer elemento es la diferencia CON SIGNO: positiva = verde arriba.
    """
    price = df[src]
    modo = modo.upper()
    kw = dict(alma_offset=alma_offset, alma_sigma=alma_sigma)

    # Price1 = sma(src,1) = src   ·   Price2 = sma(src[1],1) = src[1]
    p1, p2 = price, price.shift(1)

    if modo == "RSI":
        d = p1 - p2
        bulls = 0.5 * (d.abs() + d)
        bears = 0.5 * (d.abs() - d)
    elif modo in ("STOCHASTIC", "STOCH"):
        # El Pine usa Price1 (el cierre) para el lowest y el highest,
        # no el minimo y el maximo de la barra.
        bulls = p1 - p1.rolling(int(length), min_periods=int(length)).min()
        bears = p1.rolling(int(length), min_periods=int(length)).max() - p1
    elif modo == "ADX":
        h, l = df["High"], df["Low"]
        dh = h - h.shift(1)
        dl = l.shift(1) - l
        bulls = 0.5 * (dh.abs() + dh)
        bears = 0.5 * (dl.abs() + dl)
    else:
        raise ValueError(f"modo ASH desconocido: {modo}")

    avg_bulls = ma(ma_type, bulls, length, **kw)
    avg_bears = ma(ma_type, bears, length, **kw)
    smth_bulls = ma(ma_type, avg_bulls, smooth, **kw)
    smth_bears = ma(ma_type, avg_bears, smooth, **kw)
    return smth_bulls, smth_bears, smth_bulls - smth_bears


def ash_norm(bulls, bears):
    """
    (bulls - bears) / (bulls + bears), acotado a [-1, 1].

    El ASH crudo esta en unidades de precio: un papel de USD 500 muestra
    numeros mas grandes que uno de USD 20 aunque la fuerza sea la misma.
    El signo del crudo si sirve (es lo que pediste); esta version ademas
    permite ORDENAR el universo de mayor a menor fuerza. Las dos estan en
    el Excel.
    """
    b, s = float(bulls), float(bears)
    t = b + s
    if not np.isfinite(t) or t == 0:
        return np.nan
    return (b - s) / t


# ==============================================================================
# 4. RESTO DE INDICADORES
# ==============================================================================

def calc_rsi(c, n=14):
    d = c.diff()
    gan = rma(d.clip(lower=0), n)
    per = rma(-d.clip(upper=0), n)
    with np.errstate(divide="ignore", invalid="ignore"):
        r = 100 - 100 / (1 + gan / per)
    return r.mask((gan == 0) & (per == 0), 50.0)


def calc_atr(df, n=14):
    h, l, c = df["High"], df["Low"], df["Close"]
    pc = c.shift(1)
    tr = pd.concat([h - l, (h - pc).abs(), (l - pc).abs()], axis=1).max(axis=1)
    return rma(tr, n)


def calc_adx(df, n=14):
    h, l = df["High"], df["Low"]
    up, dn = h.diff(), -l.diff()
    plus = pd.Series(np.where((up > dn) & (up > 0), up, 0.0), index=df.index)
    minus = pd.Series(np.where((dn > up) & (dn > 0), dn, 0.0), index=df.index)
    atr = calc_atr(df, n)
    pdi = 100 * rma(plus, n) / atr
    mdi = 100 * rma(minus, n) / atr
    dx = 100 * (pdi - mdi).abs() / (pdi + mdi).replace(0, np.nan)
    return rma(dx, n)


def calc_adr_pct(df, n=20):
    """
    ADR% = 100 * (promedio de High/Low de las ultimas n ruedas - 1).

    Es el "cuanto se mueve por dia" en porcentaje, la version que usan los
    swing traders. Distinto del ATR%: el ADR ignora los gaps.
    """
    return 100 * (sma(df["High"] / df["Low"], n) - 1)


def barras_desde_cruce(hist):
    """Barras transcurridas desde el ultimo cambio de signo del histograma."""
    h = hist.dropna()
    if len(h) < 2:
        return np.nan
    signo = np.sign(h.to_numpy())
    idx = np.where(signo[1:] != signo[:-1])[0]
    if len(idx) == 0:
        return np.nan
    return int(len(signo) - 1 - (idx[-1] + 1))


def perf(c, barras):
    if len(c) <= barras:
        return np.nan
    return float(c.iloc[-1] / c.iloc[-1 - barras] - 1)


def a_semanal(df):
    """Arma barras semanales (cierre viernes) a partir de las diarias."""
    w = df.resample("W-FRI").agg({"Open": "first", "High": "max", "Low": "min",
                                  "Close": "last", "Volume": "sum"})
    return w.dropna(subset=["Close"])


# ==============================================================================
# 5. DATOS
# ==============================================================================

def cargar_mapa_cedears(path=MAPA_CEDEARS):
    """
    Mapa   ticker local  ->  simbolo del subyacente donde realmente cotiza.

    Los indicadores NUNCA se calculan sobre el CEDEAR: su precio mezcla el
    movimiento del papel con el tipo de cambio implicito, asi que el ASH, el
    RSI y el ADR saldrian contaminados. Se baja el subyacente y listo.
    """
    mapa = {}
    p = Path(path)
    if not p.exists():
        return mapa
    for linea in p.read_text(encoding="utf-8").splitlines():
        linea = linea.strip()
        if not linea or linea.startswith("#"):
            continue
        partes = [x.strip().upper() for x in linea.split(",")]
        if partes[0] in ("LOCAL", "CEDEAR", "TICKER"):
            continue
        # Una linea sin subyacente no se puede usar: antes rompia el archivo
        # entero (se colaba la clave de la linea anterior y quedaba mapeada a
        # un valor vacio o directamente reventaba con IndexError).
        if len(partes) < 2 or not partes[0] or not partes[1]:
            continue
        clave = partes[0]
        if clave.endswith(".BA"):
            clave = clave[:-3]
        mapa[clave] = partes[1]
    return mapa


def leer_universo(path, mapa=None):
    """
    universo.csv:   ticker , grupo , [subyacente]

    - Si el ticker termina en .BA se reemplaza por su subyacente, buscandolo
      primero en la tercera columna y despues en cedears.csv.
    - Si no hay forma de resolverlo, se avisa y se descarta: bajar el CEDEAR
      daria indicadores equivocados, que es peor que no tenerlos.
    """
    mapa = mapa if mapa is not None else cargar_mapa_cedears()
    p = Path(path)
    if not p.exists():
        sys.exit(f"[X] No encuentro el archivo de universo: {p}")

    filas, sin_resolver = [], []
    for linea in p.read_text(encoding="utf-8").splitlines():
        linea = linea.strip()
        if not linea or linea.startswith("#"):
            continue
        partes = [x.strip() for x in linea.split(",")]
        if partes[0].lower() in ("ticker", "simbolo", "symbol"):
            continue
        crudo = partes[0].upper()
        grupo = partes[1] if len(partes) > 1 else ""
        explicito = partes[2].upper() if len(partes) > 2 and partes[2] else ""

        local, descarga = "", crudo
        if explicito:
            local, descarga = (crudo, explicito) if explicito != crudo else ("", crudo)
        elif crudo.endswith(".BA"):
            base = crudo[:-3]
            if base in mapa:
                local, descarga = crudo, mapa[base]
            else:
                sin_resolver.append(crudo)
                continue
        filas.append({"ticker": descarga, "grupo": grupo, "local": local})

    if sin_resolver:
        print(f"    [!] sin subyacente conocido, los salteo: {', '.join(sin_resolver)}")
        print("        agregalos a cedears.csv o poné el subyacente en la 3a columna")

    df = pd.DataFrame(filas).drop_duplicates("ticker")
    if df.empty:
        sys.exit("[X] El universo quedo vacio.")
    return df


COLUMNAS_OHLCV = ["Open", "High", "Low", "Close", "Volume"]


def limpiar_barras(d, minimo=None):
    """
    Deja un DataFrame OHLCV usable, o None si no da la talla.

    Yahoo devuelve cosas raras a menudo: fechas repetidas, indice desordenado,
    filas con Close pero sin Open/High/Low, y de vez en cuando un indice con
    zona horaria que despues no se puede comparar con otro sin ella. Todo eso
    hay que emprolijarlo aca, porque mas adelante ya viaja al navegador y un
    None en High/Low le rompe el ADR y el grafico.
    """
    if d is None or len(d) == 0:
        return None
    d = d.copy()
    faltan = [c for c in COLUMNAS_OHLCV if c not in d.columns]
    if faltan:
        return None
    d = d[COLUMNAS_OHLCV].apply(pd.to_numeric, errors="coerce")

    idx = pd.to_datetime(d.index, errors="coerce")
    if getattr(idx, "tz", None) is not None:
        idx = idx.tz_convert(None)          # sin zona: se compara con cualquiera
    d.index = idx
    d = d[~d.index.isna()]
    d = d[~d.index.duplicated(keep="last")].sort_index()

    d = d.dropna(subset=["Close"])
    d = d[d["Close"] > 0]
    if len(d) == 0:
        return None
    # Una barra sin apertura o sin maximo/minimo se completa con el cierre en
    # vez de tirarse: perder la barra corre todas las ventanas moviles.
    for c in ("Open", "High", "Low"):
        d[c] = d[c].fillna(d["Close"])
    d["High"] = d[["High", "Open", "Close"]].max(axis=1)
    d["Low"] = d[["Low", "Open", "Close"]].min(axis=1)
    d["Volume"] = d["Volume"].fillna(0).clip(lower=0)
    minimo = MIN_BARRAS if minimo is None else minimo
    return d if len(d) >= minimo else None


def _descargar(grupo, periodo, minimo=None):
    """Una tanda contra Yahoo. Devuelve solo lo que vino limpio y completo."""
    import yfinance as yf
    try:
        raw = yf.download(grupo, period=periodo, interval="1d",
                          auto_adjust=True, group_by="ticker",
                          threads=True, progress=False)
    except Exception as e:
        print(f"      [!] fallo la tanda: {e}")
        return {}
    if raw is None or len(raw) == 0:
        return {}
    out = {}
    multi = isinstance(raw.columns, pd.MultiIndex)
    for t in grupo:
        try:
            if multi:
                if t not in raw.columns.get_level_values(0):
                    continue
                d = raw[t]
            else:
                d = raw
            d = limpiar_barras(d, minimo)
            if d is not None:
                out[t] = d
        except Exception:
            pass
    return out


def bajar_precios(tickers, periodo, lote=50, saltear=None, progreso=None,
                  minimo=None):
    """
    Baja los precios diarios, con tres vueltas de reintento.

    Los fallos de Yahoo vienen en RACHAS, no por simbolo: un lote entero puede
    volver vacio aunque los papeles esten perfectamente vivos (paso con BK y
    con MMC, que cotizan todos los dias). Por eso lo que no entra en la primera
    vuelta se vuelve a pedir en grupos de 5 y despues de a uno, con pausas cada
    vez mas largas. Sin esto al sitio le faltaban papeles sanos.

    `saltear` es la cuarentena: los deslistados de verdad no se piden.
    `progreso` es un callback (hechos, total, texto) para la barra del servidor.

    OJO: esta funcion la usan generar_sitio.py Y servidor.py. Estuvo duplicada
    una vez y solo una de las dos reintentaba; no la vuelvas a duplicar.
    """
    pedir = [t for t in tickers if not saltear or t not in saltear]
    if saltear:
        omitidos = len(tickers) - len(pedir)
        if omitidos:
            print(f"    {omitidos} en cuarentena, no los pido")

    datos = {}
    vueltas = [("lotes", lote, 0.6, None), ("de a 5", 5, 1.5, None),
               ("de a 1", 1, 3.0, MAX_INDIVIDUALES)]
    for vuelta, (nombre, tam, pausa, tope) in enumerate(vueltas):
        faltan = [t for t in pedir if t not in datos]
        if not faltan:
            break
        # Cortafuegos: si la primera vuelta no trajo NADA, no es una racha, es
        # que Yahoo no esta contestando. Reintentar 465 simbolos de a uno con
        # pausas seria media hora tirada y encima se come el timeout del
        # workflow. Mejor fallar rapido y publicar los datos de ayer.
        if vuelta and not datos:
            print("    [X] no vino un solo simbolo: Yahoo no esta respondiendo, "
                  "corto los reintentos")
            break
        if tope and len(faltan) > tope:
            print(f"    reintento {nombre}: {len(faltan)} pendientes, "
                  f"pruebo los primeros {tope}")
            faltan = faltan[:tope]
        elif tam != lote:
            print(f"    reintento {nombre}: {len(faltan)} pendientes", flush=True)
        for i in range(0, len(faltan), tam):
            grupo = faltan[i:i + tam]
            if tam == lote:
                print(f"    lote {i // tam + 1}: {len(grupo)} simbolos...", flush=True)
            datos.update(_descargar(grupo, periodo, minimo))
            if progreso:
                texto = ("bajando precios" if tam == lote
                         else f"reintentando los que fallaron ({nombre})")
                progreso(len(datos), len(pedir), texto)
            time.sleep(pausa)
    return datos


# ==============================================================================
# 5b. PRECIOS ATRASADOS Y CUARENTENA
# ==============================================================================
#
# El problema mas molesto del screener no es que falte un papel: es que un papel
# aparezca con el precio y la variacion de hace tres dias sin avisar. Pasa
# porque Yahoo, cuando lo apuran, devuelve la serie recortada en vez de un
# error. Se detecta comparando cada simbolo contra los de SU MISMO mercado: un
# dia de atraso en Brasil o en Europa suele ser un feriado local, no un error.

def sufijo_mercado(ticker):
    """`.SA`, `.DE`, `.TO`... Sin sufijo = Estados Unidos."""
    return ticker[ticker.rindex("."):].upper() if "." in ticker[1:] else ""


def ultima_fecha(d):
    return pd.Timestamp(d.index[-1]).normalize() if d is not None and len(d) else None


def atrasos(precios):
    """Dias habiles de atraso de cada simbolo. Ver atrasos_por_fecha."""
    ultimas = {t: ultima_fecha(d) for t, d in precios.items()}
    return atrasos_por_fecha({t: f for t, f in ultimas.items() if f is not None})


def atrasos_por_fecha(ultimas):
    """
    Dias habiles de atraso de cada simbolo contra la ultima rueda de su mercado.

    Devuelve {ticker: dias}. 0 = al dia. Se compara por mercado y no contra la
    fecha maxima global para no marcar como atrasado a media Europa cada vez
    que tienen un feriado propio.
    """
    if not ultimas:
        return {}
    refs = {}
    for t, f in ultimas.items():
        s = sufijo_mercado(t)
        refs.setdefault(s, []).append(f)
    # la referencia del mercado es la fecha mas frecuente, no la maxima: un solo
    # simbolo adelantado (o con una barra basura) no puede correr la vara.
    ref = {}
    for s, fechas in refs.items():
        vc = pd.Series(fechas).value_counts()
        top = vc.max()
        ref[s] = max(f for f, n in vc.items() if n == top)
    out = {}
    for t, f in ultimas.items():
        r = ref[sufijo_mercado(t)]
        out[t] = max(0, int(np.busday_count(f.date(), r.date()))) if f < r else 0
    return out


def repescar_atrasados(precios, periodo, maximo=80, progreso=None):
    """
    Vuelve a pedir de a uno los simbolos que quedaron atrasados.

    Cuando un lote grande viene recortado, pedir el simbolo solo casi siempre
    devuelve la serie completa. Se hace despues de la descarga y solo con los
    atrasados, asi que cuesta poco.
    """
    tarde = [t for t, n in atrasos(precios).items() if n > 0]
    if not tarde:
        return precios, []
    tarde = sorted(tarde, key=lambda t: -atrasos(precios)[t])[:maximo]
    print(f"    {len(tarde)} con la ultima barra atrasada, los repesco de a uno...")
    for k, t in enumerate(tarde, 1):
        if progreso:
            progreso(k, len(tarde), "repescando los que vinieron atrasados")
        nuevo = _descargar([t], periodo).get(t)
        if nuevo is not None:
            viejo = precios.get(t)
            if viejo is None or ultima_fecha(nuevo) >= ultima_fecha(viejo):
                precios[t] = nuevo
        time.sleep(0.8)
    quedan = [t for t, n in atrasos(precios).items() if n > 0]
    if quedan:
        print(f"    siguen atrasados ({len(quedan)}): {', '.join(sorted(quedan)[:20])}"
              + (" ..." if len(quedan) > 20 else ""))
    return precios, quedan


def cargar_cuarentena(path=CUARENTENA):
    """{ticker: {"fallos": n, "hasta": iso}} de lo que Yahoo no devuelve."""
    p = Path(path)
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def simbolos_en_cuarentena(cuarentena):
    """Los que todavia no cumplieron la semana de castigo."""
    hoy = datetime.now(timezone.utc).date().isoformat()
    return {t for t, v in cuarentena.items()
            if v.get("fallos", 0) >= CUARENTENA_FALLOS and v.get("hasta", "") > hoy}


def actualizar_cuarentena(cuarentena, pedidos, obtenidos, path=CUARENTENA):
    """
    Suma un fallo a lo que no vino y perdona a lo que si vino.

    A los tres fallos seguidos el simbolo queda afuera una semana. Pasada la
    semana se vuelve a pedir solo: los deslistados de verdad (WBA, TTM, LFC...)
    reinciden y los que fallaron por una racha de Yahoo vuelven a entrar.
    """
    hasta = (datetime.now(timezone.utc) + pd.Timedelta(days=CUARENTENA_DIAS)).date().isoformat()
    for t in pedidos:
        if t in obtenidos:
            cuarentena.pop(t, None)
        else:
            v = cuarentena.get(t, {"fallos": 0})
            v["fallos"] = v.get("fallos", 0) + 1
            v["hasta"] = hasta
            cuarentena[t] = v
    try:
        Path(path).write_text(json.dumps(cuarentena, indent=1), encoding="utf-8")
    except Exception as e:
        print(f"    [!] no pude guardar la cuarentena: {e}")
    return cuarentena


def bajar_fundamentales(tickers, usar_cache=True):
    import yfinance as yf
    cache = {}
    if usar_cache and CACHE_META.exists():
        try:
            cache = json.loads(CACHE_META.read_text())
        except Exception:
            cache = {}
    ahora = datetime.now(timezone.utc).timestamp()
    faltan = [t for t in tickers
              if t not in cache or ahora - cache[t].get("_ts", 0) > CACHE_DIAS * 86400]
    print(f"    {len(tickers) - len(faltan)} en cache, {len(faltan)} a descargar")

    def uno(t):
        try:
            i = yf.Ticker(t).get_info()
            return t, {"nombre": i.get("shortName") or i.get("longName") or "",
                       "sector": i.get("sector") or "",
                       "industria": i.get("industry") or "",
                       "pais": i.get("country") or "",
                       "float_shares": i.get("floatShares"),
                       "shares_out": i.get("sharesOutstanding"),
                       "mcap": i.get("marketCap"), "_ts": ahora}
        except Exception:
            return t, {"nombre": "", "sector": "", "industria": "", "pais": "",
                       "float_shares": None, "shares_out": None, "mcap": None,
                       "_ts": ahora}

    if faltan:
        with ThreadPoolExecutor(max_workers=6) as ex:
            for n, (t, d) in enumerate(ex.map(uno, faltan), 1):
                cache[t] = d
                if n % 50 == 0:
                    print(f"      {n}/{len(faltan)}", flush=True)
        try:
            CACHE_META.write_text(json.dumps(cache))
        except Exception as e:
            print(f"    [!] no pude guardar el cache: {e}")
    return cache


def guardar_precios(datos, meta=None, path=CACHE_PRECIOS):
    """Deja los precios en disco para que la app web no vuelva a bajarlos."""
    import pickle
    with open(path, "wb") as fh:
        pickle.dump({"precios": datos, "meta": meta or {},
                     "fecha": datetime.now().isoformat(timespec="minutes")}, fh)


def cargar_precios(path=CACHE_PRECIOS):
    """Devuelve (precios, meta, fecha) o (None, None, None) si no hay cache."""
    import pickle
    p = Path(path)
    if not p.exists():
        return None, None, None
    try:
        with open(p, "rb") as fh:
            d = pickle.load(fh)
        return d["precios"], d.get("meta", {}), d.get("fecha", "")
    except Exception:
        return None, None, None


# ==============================================================================
# 6. METRICAS
# ==============================================================================

def metricas(t, df, meta, bench_perf, cfg_ash=None, paragon=None, adr_len=None,
             rsi_len=None, atr_len=None, adx_len=None, historial=0):
    """
    Calcula todas las metricas de un simbolo.

    Los parametros vienen por argumento (no de las constantes) para que la app
    web pueda recalcular con la configuracion que el usuario elija sin tocar
    el modulo. Si no se pasan, se usan los defaults de arriba.

    historial: si es > 0, agrega las ultimas N barras del ASH como lista, para
    dibujar el sparkline en la tabla.
    """
    cfg = dict(cfg_ash or CFG_ASH)
    par = dict(PARAGON, **(paragon or {}))
    adr_len = adr_len or ADR_LEN
    rsi_len = rsi_len or RSI_LEN
    atr_len = atr_len or ATR_LEN
    adx_len = adx_len or ADX_LEN

    c, v = df["Close"], df["Volume"]
    px = float(c.iloc[-1])

    # --- ASH diario ---
    bu, be, ash_d = calc_ash(df, **cfg)

    # --- ASH semanal (barras armadas desde las diarias) ---
    sem = a_semanal(df)
    ash_w = None
    if len(sem) >= MIN_BARRAS_SEM:
        buw, bew, ash_w = calc_ash(sem, **cfg)
        w_val = float(ash_w.iloc[-1])
        w_prev = float(ash_w.iloc[-2]) if len(ash_w) >= 2 else np.nan
        w_norm = ash_norm(buw.iloc[-1], bew.iloc[-1])
    else:
        w_val = w_prev = w_norm = np.nan

    vol20 = float(v.rolling(20).mean().iloc[-1])
    vent = min(len(c), 252)
    max52, min52 = float(c.iloc[-vent:].max()), float(c.iloc[-vent:].min())

    fs = meta.get("float_shares") or meta.get("shares_out")
    fs = float(fs) if fs else np.nan
    mcap = meta.get("mcap")
    mcap = float(mcap) if mcap else np.nan

    d_val = float(ash_d.iloc[-1])
    d_prev = float(ash_d.iloc[-2]) if len(ash_d) >= 2 else np.nan
    p3m = perf(c, 63)

    f = {
        "ticker": t,
        "nombre": meta.get("nombre", ""),
        "sector": meta.get("sector", ""),
        "industria": meta.get("industria", ""),
        "pais": meta.get("pais", ""),
        "precio": px,
        "chg_pct": perf(c, 1),

        "ash_d": d_val,
        "ash_d_pend": d_val - d_prev,
        "ash_d_norm": ash_norm(bu.iloc[-1], be.iloc[-1]),
        "ash_d_positivo": bool(d_val > 0),
        "ash_d_creciendo": bool(d_val > d_prev),
        "ash_d_barras_cruce": barras_desde_cruce(ash_d),
        "ash_bulls_d": float(bu.iloc[-1]),
        "ash_bears_d": float(be.iloc[-1]),

        "ash_w": w_val,
        "ash_w_pend": w_val - w_prev,
        "ash_w_norm": w_norm,
        "ash_w_positivo": bool(w_val > 0) if w_val == w_val else False,
        "ash_w_creciendo": bool(w_val > w_prev) if w_val == w_val else False,

        "rsi14": float(calc_rsi(c, rsi_len).iloc[-1]),
        "adr_pct": float(calc_adr_pct(df, adr_len).iloc[-1]),
        "atr_pct": float(calc_atr(df, atr_len).iloc[-1] / px * 100),
        "adx14": float(calc_adx(df, adx_len).iloc[-1]),

        "perf_1s": perf(c, 5), "perf_1m": perf(c, 21), "perf_3m": p3m,
        "perf_6m": perf(c, 126), "perf_12m": perf(c, 252),
        "rs_3m": (p3m - bench_perf) if (p3m == p3m and bench_perf == bench_perf) else np.nan,
        "dist_max52s": px / max52 - 1,
        "dist_min52s": px / min52 - 1,

        "volumen": float(v.iloc[-1]),
        "vol20": vol20,
        "rel_vol": float(v.iloc[-1] / vol20) if vol20 else np.nan,
        "dolar_vol20": float((c * v).rolling(20).mean().iloc[-1]),
        "float_shares_m": fs / 1e6 if fs == fs else np.nan,
        "float_musd": fs * px / 1e6 if fs == fs else np.nan,
        "rotacion_float": vol20 / fs if (fs == fs and fs) else np.nan,
        "mcap_musd": mcap / 1e6 if mcap == mcap else np.nan,
    }
    # --- PARAGON ---
    # Conjunto B (ancla 1D): EXACTO, son la EMA 100/200 sobre las diarias que
    # ya baja el screener. Conjunto A (ancla 4h): APROXIMADO, porque no hay
    # velas de 4h en el pipeline; se convierten las longitudes conservando la
    # tasa de decaimiento y la fila queda marcada.
    atr_v = float(calc_atr(df, 14).iloc[-1])
    rb, lb, _, _ = paragon_conjunto(c, par["rapida"], par["lenta"], 1)
    ra, la, kra, kla = paragon_conjunto(c, par["rapida"], par["lenta"], par["k"])
    B = senales_paragon(df, rb, lb, atr_v)
    A = senales_paragon(df, ra, la, atr_v)
    for pre, S in (("par_b", B), ("par_a", A)):
        f[f"{pre}_sesgo"] = S["sesgo"]
        f[f"{pre}_pos"] = S["pos"]
        f[f"{pre}_ancho"] = S["ancho"]
        f[f"{pre}_dist"] = S["dist"]
        f[f"{pre}_dist_atr"] = S["dist_atr"]
        f[f"{pre}_cruce"] = S["cruce"]
        f[f"{pre}_fresco"] = bool(S["cruce"] == S["cruce"]
                                  and S["cruce"] <= par["fresco"])
    f["par_a_aprox"] = par["k"] != 1
    f["par_a_largos"] = f"{kra}/{kla}"
    if A["sesgo"] is None or B["sesgo"] is None:
        f["regimen"], f["regimen_ord"] = "", np.nan
    else:
        f["regimen"] = ("A+ B+" if (A["sesgo"] and B["sesgo"]) else
                        "A− B+" if B["sesgo"] else
                        "A+ B−" if A["sesgo"] else "A− B−")
        f["regimen_ord"] = (2 if B["sesgo"] else 0) + (1 if A["sesgo"] else 0)
    rv, llena = rvwap_expansivo(df, par["rv_len"], par["rv_fuente"])
    f["rvwap"] = float(rv.iloc[-1]) if np.isfinite(rv.iloc[-1]) else np.nan
    f["vs_rvwap"] = px / f["rvwap"] - 1 if f["rvwap"] == f["rvwap"] else np.nan
    f["rv_llena"] = bool(llena)
    # los que no llegan al warmup no se descartan en silencio
    f["sin_historial"] = bool(A["sesgo"] is None or B["sesgo"] is None)
    if historial:
        f["serie_d"] = [float(x) for x in ash_d.dropna().iloc[-historial:]]
        f["serie_w"] = ([float(x) for x in ash_w.dropna().iloc[-historial:]]
                        if ash_w is not None else [])
    return f


def aplicar_filtros(df, filtros):
    m = pd.Series(True, index=df.index)
    for clave, col, op in REGLAS:
        val = filtros.get(clave)
        if val is None or col not in df.columns:
            continue
        if op == "ge":
            m &= df[col] >= val
        elif op == "le":
            m &= df[col] <= val
        elif op == "bool":
            m &= df[col].astype(bool) == bool(val)

    v = filtros.get("ash_d_barras_cruce_max")
    if v is not None:
        m &= (df["ash_d_barras_cruce"] <= v) & (df["ash_d"] > 0)
    v = filtros.get("dist_max52s_max")
    if v is not None:
        m &= df["dist_max52s"] >= -abs(v) / 100
    return m.fillna(False)


# ==============================================================================
# 7. EXCEL
# ==============================================================================

def exportar(df_todo, df_filtrado, salida, notas):
    from openpyxl import Workbook
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    fuente = "Arial"
    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(name=fuente, bold=True, color="FFFFFF", size=10)
    body_font = Font(name=fuente, size=10)
    borde = Border(bottom=Side(style="thin", color="D9D9D9"))
    verde = Font(name=fuente, size=10, color="1E7B34", bold=True)
    rojo = Font(name=fuente, size=10, color="B02418", bold=True)

    def hoja(nombre, datos):
        ws = wb.create_sheet(nombre)
        claves = [k for k, _, _ in COLUMNAS if k in datos.columns]
        titulos = {k: h for k, h, _ in COLUMNAS}
        formatos = {k: f for k, _, f in COLUMNAS}
        for j, k in enumerate(claves, 1):
            cl = ws.cell(row=1, column=j, value=titulos[k])
            cl.fill, cl.font = head_fill, head_font
            cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for i, (_, r) in enumerate(datos.iterrows(), 2):
            for j, k in enumerate(claves, 1):
                v = r[k]
                if isinstance(v, (bool, np.bool_)):
                    v = "SI" if v else "NO"
                elif isinstance(v, (float, np.floating)):
                    v = None if pd.isna(v) else float(v)
                elif isinstance(v, (int, np.integer)):
                    v = int(v)
                cel = ws.cell(row=i, column=j, value=v)
                cel.border = borde
                cel.number_format = formatos[k]
                if k in ("ash_d", "ash_w") and isinstance(v, float):
                    cel.font = verde if v > 0 else rojo
                else:
                    cel.font = body_font
        n = len(datos) + 1
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(claves))}{n}"
        ws.row_dimensions[1].height = 32
        anchos = {"nombre": 26, "industria": 26, "sector": 20, "pais": 14,
                  "ticker": 10, "grupo": 14}
        for j, k in enumerate(claves, 1):
            ws.column_dimensions[get_column_letter(j)].width = anchos.get(k, 12)
        for k in ("ash_d_norm", "ash_w_norm", "rs_rank", "adr_pct"):
            if k in claves and n > 1:
                col = get_column_letter(claves.index(k) + 1)
                ws.conditional_formatting.add(
                    f"{col}2:{col}{n}",
                    ColorScaleRule(start_type="min", start_color="F8696B",
                                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                                   end_type="max", end_color="63BE7B"))
        return ws

    hoja("Filtrado", df_filtrado)
    hoja("Universo", df_todo)
    ws = wb.create_sheet("Config")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 50
    for i, (a, b) in enumerate(notas, 1):
        ws.cell(row=i, column=1, value=a).font = Font(name=fuente, bold=True, size=10)
        ws.cell(row=i, column=2, value=str(b)).font = Font(name=fuente, size=10)
    wb.save(salida)


# ==============================================================================
# 8. MAIN
# ==============================================================================

def main():
    ap = argparse.ArgumentParser(description="Screener local con ASH diario y semanal")
    ap.add_argument("--universo", default="universo.csv")
    ap.add_argument("--out", default="screener.xlsx")
    ap.add_argument("--periodo", default=PERIODO)
    ap.add_argument("--sin-fundamentales", action="store_true")
    ap.add_argument("--usar-cache", action="store_true",
                    help="usa los precios ya bajados (cache_precios.pkl)")
    args = ap.parse_args()

    t0 = time.time()
    uni = leer_universo(args.universo)
    tickers = uni["ticker"].tolist()
    grupos = dict(zip(uni["ticker"], uni["grupo"]))
    locales = dict(zip(uni["ticker"], uni["local"]))
    print(f"[1/5] Universo: {len(tickers)} simbolos")

    print("[2/5] Precios diarios...")
    pedir = tickers if BENCHMARK in tickers else tickers + [BENCHMARK]
    if args.usar_cache:
        precios, _, fecha = cargar_precios()
        if precios is None:
            sys.exit("[X] No hay cache de precios. Corré sin --usar-cache primero.")
        print(f"      cache del {fecha}")
    else:
        precios = bajar_precios(pedir, args.periodo)
    bench = precios.get(BENCHMARK)
    if BENCHMARK not in tickers:
        precios.pop(BENCHMARK, None)
    bench_perf = perf(bench["Close"], 63) if bench is not None else np.nan
    faltantes = [t for t in tickers if t not in precios]
    print(f"      {len(precios)} con datos, {len(faltantes)} descartados")

    if not args.usar_cache:
        guardar_precios(precios)

    print("[3/5] Fundamentales...")
    meta = ({t: {} for t in precios} if args.sin_fundamentales
            else bajar_fundamentales(list(precios.keys())))

    print("[4/5] Indicadores (ASH D + ASH W, EMAs, ADR, RSI)...")
    filas = []
    for t, d in precios.items():
        try:
            f = metricas(t, d, meta.get(t, {}), bench_perf)
            f["grupo"] = grupos.get(t, "")
            f["local"] = locales.get(t, "")
            filas.append(f)
        except Exception as e:
            print(f"      [!] {t}: {e}")
    df = pd.DataFrame(filas)
    if df.empty:
        sys.exit("[X] Ningun simbolo con datos. Revisa conexion, tickers o MIN_BARRAS.")

    df["rs_rank"] = (df["perf_3m"].rank(pct=True) * 98 + 1).round()
    mask = aplicar_filtros(df, FILTROS)
    df = df.sort_values("ash_d_norm", ascending=False)
    df_f = df[mask.reindex(df.index, fill_value=False)]
    print(f"      {len(df_f)} de {len(df)} pasan el filtro")

    print("[5/5] Excel...")
    notas = [("Generado", datetime.now().strftime("%Y-%m-%d %H:%M")),
             ("Universo", f"{len(tickers)} simbolos ({args.universo})"),
             ("Con datos", len(df)), ("Pasan el filtro", len(df_f)),
             ("Descartados", ", ".join(faltantes) if faltantes else "ninguno"),
             ("Periodo", args.periodo), ("Benchmark", BENCHMARK), ("", ""),
             ("--- ASH ---", "Absolute Strength Histogram v2 | jh"),
             ("columna ASH", "SmthBulls - SmthBears (con signo)"),
             ("modo", CFG_ASH["modo"]), ("length", CFG_ASH["length"]),
             ("smooth", CFG_ASH["smooth"]), ("ma_type", CFG_ASH["ma_type"]),
             ("semanal", "barras W-FRI armadas desde las diarias"), ("", ""),
             ("--- Otros ---", ""),
             ("Paragon", f'{PARAGON["rapida"]}/{PARAGON["lenta"]} · k={PARAGON["k"]}'),
             ("ADR", f"{ADR_LEN} ruedas"), ("RSI", RSI_LEN), ("", ""),
             ("--- Filtros activos ---", "")] + \
            [(k, v) for k, v in FILTROS.items() if v is not None]
    exportar(df, df_f, args.out, notas)
    print(f"\nListo -> {args.out}   ({time.time() - t0:.0f}s)")


if __name__ == "__main__":
    main()


# ==============================================================================
# 9. PRE-MARKET Y AFTER-HOURS
# ==============================================================================
#
# DE DONDE SALEN. Del mismo endpoint de graficos de Yahoo, pero con dos cosas
# distintas: intervalo intradia (5m) y includePrePost=1. Ahi vienen las barras
# de las tres sesiones pegadas. yfinance lo expone como prepost=True.
#
# COMO SE SEPARAN. Yahoo no marca cada barra con su sesion, asi que hay que
# mirar la hora local del mercado: antes de las 9:30 es pre-market, de 16:00 en
# adelante es after-hours. yfinance ya devuelve el indice en la zona horaria del
# mercado, asi que no hay que adivinar husos.
#
# CONTRA QUE SE COMPARA. El pre-market se compara contra el cierre del dia
# ANTERIOR, y el after-hours contra el cierre de HOY. Compararlos contra
# cualquier otra cosa da porcentajes que no significan nada.
#
# LO QUE NO SE PUEDE PROMETER. Fuera de horario el volumen es una fraccion del
# de la rueda: en un papel liquido el dato es util, en uno que no lo es puede
# haber una sola operacion suelta moviendo el precio 4%. Por eso viaja tambien
# el volumen extendido: sirve para saber cuanto creerle.

APERTURA_REGULAR = (9, 30)     # hora local del mercado
CIERRE_REGULAR = (16, 0)


def _minutos(ts):
    return ts.hour * 60 + ts.minute


def extendido_de_barras(d, cierre_previo=None):
    """
    Saca el precio de pre-market o after-hours de un DataFrame de barras 5m.

    Devuelve {"px":, "pct":, "tipo": "pre"|"post", "vol":} o None si la ultima
    barra es de la rueda regular (o si no hay nada aprovechable).
    """
    if d is None or len(d) == 0 or "Close" not in d:
        return None
    d = d.dropna(subset=["Close"])
    if len(d) == 0:
        return None
    ini = APERTURA_REGULAR[0] * 60 + APERTURA_REGULAR[1]
    fin = CIERRE_REGULAR[0] * 60 + CIERRE_REGULAR[1]

    idx = d.index
    if getattr(idx, "tz", None) is None:
        return None                      # sin zona no se puede saber la sesion
    minutos = np.array([_minutos(t) for t in idx])
    regular = (minutos >= ini) & (minutos < fin)

    ult = len(d) - 1
    if regular[ult]:
        return None                      # la rueda esta abierta: no hay extendido
    tipo = "pre" if minutos[ult] < ini else "post"

    # referencia: el ultimo cierre regular que haya en la ventana; si no hay
    # (pre-market de un dia nuevo), el cierre diario que ya teniamos
    ref = None
    if regular.any():
        ref = float(d["Close"].to_numpy()[regular][-1])
    if tipo == "pre" or ref is None:
        ref = float(cierre_previo) if cierre_previo else ref
    if not ref or not np.isfinite(ref):
        return None

    px = float(d["Close"].iloc[-1])
    fuera = ~regular
    vol = float(np.nansum(d["Volume"].to_numpy()[fuera])) if "Volume" in d else 0.0
    return {"px": round(px, 4), "pct": round(px / ref - 1, 6),
            "tipo": tipo, "vol": int(vol)}


def bajar_extendido(tickers, cierres=None, lote=50, progreso=None):
    """
    {ticker: {"px","pct","tipo","vol"}} para los que tengan pre o after.

    SOLO PARA ESTADOS UNIDOS. La ventana 9:30-16:00 que usa extendido_de_barras
    es la de Nueva York, y aplicarsela a Tokio o a Sao Paulo da cualquier cosa:
    en la primera corrida real los dos unicos "pre-market" que salieron fueron
    6701.T y 2317.TW, justamente por eso. Ademas Yahoo casi no publica sesiones
    extendidas fuera de EE.UU. Los simbolos con sufijo de mercado se saltean.

    Es una pasada aparte de la de precios diarios porque necesita otro
    intervalo. Se piden barras de 5 minutos del ultimo dia: es lo mas barato
    que cubre las tres sesiones.
    """
    import yfinance as yf
    cierres = cierres or {}
    out = {}
    tickers = [t for t in tickers if not sufijo_mercado(t)]
    for i in range(0, len(tickers), lote):
        grupo = tickers[i:i + lote]
        try:
            raw = yf.download(grupo, period="1d", interval="5m", prepost=True,
                              group_by="ticker", threads=True, progress=False,
                              auto_adjust=False)
        except Exception as e:
            print(f"      [!] fallo el lote extendido: {e}")
            continue
        if raw is None or len(raw) == 0:
            continue
        multi = isinstance(raw.columns, pd.MultiIndex)
        for t in grupo:
            try:
                if multi:
                    if t not in raw.columns.get_level_values(0):
                        continue
                    d = raw[t]
                else:
                    d = raw
                e = extendido_de_barras(d, cierres.get(t))
                if e:
                    out[t] = e
            except Exception:
                pass
        if progreso:
            progreso(len(out), len(tickers), "pre-market / after-hours")
        time.sleep(0.5)
    return out
